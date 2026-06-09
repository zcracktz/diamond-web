"""Laporan Detail Penghimpunan dan Pengolahan Data views."""

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.http import require_http_methods
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_protect
from django.db.models import Q
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..models.ilap import ILAP
from ..models.jenis_data_ilap import JenisDataILAP
from ..models.klasifikasi_jenis_data import KlasifikasiJenisData


def is_pmde_user(user):
    """Check if user belongs to PMDE group."""
    return user.is_superuser or user.is_staff or user.groups.filter(
        name__in=['user_pmde', 'admin', 'admin_pmde']
    ).exists()


def _get_filtered_detail_data(params):
    """Utility function to get filtered detail data based on request parameters."""
    kategori_ilap = params.get('kategori_ilap')
    nama_ilap = params.get('nama_ilap')
    dasar_hukum = params.get('dasar_hukum')
    jenis_data = params.get('jenis_data')
    periode = params.get('periode')
    nama_tabel = params.get('nama_tabel')

    # Base query
    jenis_data_ilap_list = JenisDataILAP.objects.all().select_related(
        'id_ilap',
        'id_ilap__id_kategori_ilap',
        'id_klasifikasi_jenis_data',
        'id_dasar_hukum',
        'id_jenis_tabel'
    )

    # Filter by kategori ILAP
    if kategori_ilap and kategori_ilap != 'all' and kategori_ilap != '':
        jenis_data_ilap_list = jenis_data_ilap_list.filter(id_ilap__id_kategori_ilap_id=kategori_ilap)

    # Filter by nama ILAP
    if nama_ilap and nama_ilap != 'all' and nama_ilap != '':
        jenis_data_ilap_list = jenis_data_ilap_list.filter(id_ilap_id=nama_ilap)

    # Filter by dasar hukum
    if dasar_hukum and dasar_hukum != 'all' and dasar_hukum != '':
        jenis_data_ilap_list = jenis_data_ilap_list.filter(id_dasar_hukum_id=dasar_hukum)

    # Filter by jenis data
    if jenis_data and jenis_data != 'all' and jenis_data != '':
        jenis_data_ilap_list = jenis_data_ilap_list.filter(id=jenis_data)

    # Filter by periode pengiriman
    if periode and periode != 'all' and periode != '':
        jenis_data_ilap_list = jenis_data_ilap_list.filter(periode_jenis_data__id_periode_pengiriman_id=periode).distinct()

    # Filter by nama tabel
    if nama_tabel and nama_tabel != 'all' and nama_tabel != '':
        jenis_data_ilap_list = jenis_data_ilap_list.filter(id_jenis_tabel_id=nama_tabel)

    return jenis_data_ilap_list


class LaporanDetailHimpunOlahDataView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Display Laporan Detail Penghimpunan dan Pengolahan Data."""
    template_name = 'laporan_detail_himpun_olah_data/list.html'

    def test_func(self):
        """Allow access only to PMDE users."""
        return is_pmde_user(self.request.user)

    def get_context_data(self, **kwargs):
        """Add context data."""
        context = super().get_context_data(**kwargs)
        return context


@login_required
@user_passes_test(is_pmde_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def laporan_detail_himpun_olah_data_data(request):
    """DataTables server-side endpoint for Laporan Detail Penghimpunan dan Pengolahan Data."""
    params = request.POST if request.method == 'POST' else request.GET

    try:
        draw = int(params.get('draw', 1))
    except (ValueError, TypeError):
        draw = 1

    try:
        start = int(params.get('start', 0))
    except (ValueError, TypeError):
        start = 0

    try:
        length = int(params.get('length', 10))
    except (ValueError, TypeError):
        length = 10

    # Get filtered JenisDataILAP
    jenis_data_ilap_list = _get_filtered_detail_data(params)

    records_total = JenisDataILAP.objects.count()
    records_filtered = jenis_data_ilap_list.count()

    # Pagination
    jenis_data_paginated = jenis_data_ilap_list[start:start + length]

    # Build response data
    data = []
    for idx, jenis_data in enumerate(jenis_data_paginated, start=start + 1):
        row = {
            'kategori_ilap': jenis_data.id_ilap.id_kategori_ilap.nama_kategori_ilap if jenis_data.id_ilap and jenis_data.id_ilap.id_kategori_ilap else '',
            'nama_ilap': jenis_data.id_ilap.nama_ilap if jenis_data.id_ilap else '',
            'nama_jenis_data': jenis_data.nama_jenis_data,
            'nama_sub_jenis_data': jenis_data.nama_sub_jenis_data,
            'nama_tabel': jenis_data.id_jenis_tabel.nama_jenis_tabel if jenis_data.id_jenis_tabel else '',
            'klasifikasi': jenis_data.id_klasifikasi_jenis_data.nama_klasifikasi if jenis_data.id_klasifikasi_jenis_data else '',
            'dasar_hukum': jenis_data.id_dasar_hukum.nama_dasar_hukum if jenis_data.id_dasar_hukum else '',
            'periode_pengiriman': ', '.join([p.nama_periode_pengiriman for p in jenis_data.periode_jenis_data.filter(id_periode_pengiriman__isnull=False).values_list('id_periode_pengiriman__nama_periode_pengiriman', flat=True).distinct()]) if hasattr(jenis_data, 'periode_jenis_data') else '',
        }
        data.append(row)

    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
@user_passes_test(is_pmde_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def laporan_detail_himpun_olah_data_export(request):
    """Export Laporan Detail Penghimpunan dan Pengolahan Data to XLSX or PDF."""
    params = request.GET if request.method == 'GET' else request.POST
    export_format = params.get('format', 'excel')

    # Get filtered data
    jenis_data_ilap_list = _get_filtered_detail_data(params)

    if export_format == 'excel':
        return _export_detail_to_excel(jenis_data_ilap_list)
    elif export_format == 'pdf':
        return _export_detail_to_pdf(jenis_data_ilap_list)
    else:
        return HttpResponse('Format not supported', status=400)


def _export_detail_to_excel(jenis_data_ilap_list):
    """Export detail data to Excel format."""
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail Himpun & Olah Data"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Write title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = 'DETAIL PENGHIMPUNAN DAN PENGOLAHAN DATA'
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = center_alignment

    # Write subtitle
    ws.merge_cells('A2:I2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = 'Informasi Detail IPC Penghimpunan Data yang telah disampaikan ke AP'
    subtitle_cell.font = Font(italic=True, size=10)
    subtitle_cell.alignment = center_alignment

    # Write headers
    headers = ['NO', 'KATEGORI ILAP', 'NAMA ILAP', 'NAMA JENIS DATA', 'NAMA SUB JENIS DATA', 'NAMA TABEL', 'KLASIFIKASI', 'DASAR HUKUM', 'PERIODE PENGIRIMAN']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Write data
    for row_idx, jenis_data in enumerate(jenis_data_ilap_list, start=5):
        periode_list = jenis_data.periode_jenis_data.filter(id_periode_pengiriman__isnull=False).values_list('id_periode_pengiriman__nama_periode_pengiriman', flat=True).distinct()
        periode_str = ', '.join(periode_list) if periode_list else ''

        row_data = [
            row_idx - 4,
            jenis_data.id_ilap.id_kategori_ilap.nama_kategori_ilap if jenis_data.id_ilap and jenis_data.id_ilap.id_kategori_ilap else '',
            jenis_data.id_ilap.nama_ilap if jenis_data.id_ilap else '',
            jenis_data.nama_jenis_data,
            jenis_data.nama_sub_jenis_data,
            jenis_data.id_jenis_tabel.nama_jenis_tabel if jenis_data.id_jenis_tabel else '',
            jenis_data.id_klasifikasi_jenis_data.nama_klasifikasi if jenis_data.id_klasifikasi_jenis_data else '',
            jenis_data.id_dasar_hukum.nama_dasar_hukum if jenis_data.id_dasar_hukum else '',
            periode_str,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx == 1:
                cell.alignment = center_alignment

    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 20

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_data = excel_file.getvalue()

    # Create response
    filename = f"Detail_Penghimpunan_Pengolahan_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_detail_to_pdf(jenis_data_ilap_list):
    """Export detail data to PDF format."""
    # For now, return Excel
    return _export_detail_to_excel(jenis_data_ilap_list)
