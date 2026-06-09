"""Laporan Rekap Penghimpunan dan Pengolahan Data views."""

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.http import require_http_methods
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_protect
from django.db.models import Q, Count
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..models.tiket import Tiket
from ..models.ilap import ILAP
from ..models.jenis_data_ilap import JenisDataILAP
from ..models.klasifikasi_jenis_data import KlasifikasiJenisData
from ..models.periode_pengiriman import PeriodePengiriman
from ..models.jenis_tabel import JenisTabel
from ..models.dasar_hukum import DasarHukum


def is_pmde_user(user):
    """Check if user belongs to PMDE group."""
    return user.is_superuser or user.is_staff or user.groups.filter(
        name__in=['user_pmde', 'admin', 'admin_pmde']
    ).exists()


def _get_filtered_data(params):
    """Utility function to get filtered data based on request parameters."""
    kategori_ilap = params.get('kategori_ilap')
    nama_ilap = params.get('nama_ilap')
    dasar_hukum = params.get('dasar_hukum')
    jenis_data = params.get('jenis_data')
    periode = params.get('periode')
    nama_tabel = params.get('nama_tabel')

    # Base query
    ilaps = ILAP.objects.all()

    # Filter by kategori ILAP
    if kategori_ilap and kategori_ilap != 'all' and kategori_ilap != '':
        ilaps = ilaps.filter(id_kategori_ilap_id=kategori_ilap)

    # Filter by nama ILAP
    if nama_ilap and nama_ilap != 'all' and nama_ilap != '':
        ilaps = ilaps.filter(id=nama_ilap)

    # Filter by dasar hukum
    if dasar_hukum and dasar_hukum != 'all' and dasar_hukum != '':
        ilaps = ilaps.filter(jenis_data_ilap__id_dasar_hukum_id=dasar_hukum).distinct()

    # Filter by jenis data
    if jenis_data and jenis_data != 'all' and jenis_data != '':
        ilaps = ilaps.filter(jenis_data_ilap__id=jenis_data).distinct()

    # Filter by periode pengiriman
    if periode and periode != 'all' and periode != '':
        ilaps = ilaps.filter(periode_jenis_data__id_periode_pengiriman_id=periode).distinct()

    # Filter by nama tabel
    if nama_tabel and nama_tabel != 'all' and nama_tabel != '':
        ilaps = ilaps.filter(jenis_data_ilap__id_jenis_tabel_id=nama_tabel).distinct()

    return ilaps


class LaporanRekapHimpunOlahDataView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Display Laporan Rekap Penghimpunan dan Pengolahan Data."""
    template_name = 'laporan_rekap_himpun_olah_data/list.html'

    def test_func(self):
        """Allow access only to PMDE users."""
        return is_pmde_user(self.request.user)

    def get_context_data(self, **kwargs):
        """Add context data."""
        context = super().get_context_data(**kwargs)
        return context


@login_required
@user_passes_test(is_pmde_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def laporan_rekap_himpun_olah_data_data(request):
    """DataTables server-side endpoint for Laporan Rekap Penghimpunan dan Pengolahan Data."""
    params = request.POST if request.method == 'POST' else request.GET

    try:
        draw = int(params.get('draw', 1))
    except (ValueError, TypeError):
        draw = 1

    try:
        start = int(params.get('start', 0))
    except (ValueError, TypeError):
        start = 0

    try:
        length = int(params.get('length', 10))
    except (ValueError, TypeError):
        length = 10

    # Get filtered ILAPs
    ilaps = _get_filtered_data(params)

    records_total = ILAP.objects.count()
    records_filtered = ilaps.count()

    # Pagination
    ilaps_paginated = ilaps[start:start + length]

    # Build response data
    data = []
    for idx, ilap in enumerate(ilaps_paginated, start=start + 1):
        # Get jenis data for this ILAP
        jenis_data_list = ilap.jenis_data_ilap.all()

        # Count by klasifikasi
        wajib_count = jenis_data_list.filter(
            id_klasifikasi_jenis_data__nama_klasifikasi='WAJIB'
        ).count()
        penting_count = jenis_data_list.filter(
            id_klasifikasi_jenis_data__nama_klasifikasi='PENTING'
        ).count()
        lengkap_count = jenis_data_list.filter(
            id_klasifikasi_jenis_data__nama_klasifikasi='LENGKAP'
        ).count()
        langka_count = jenis_data_list.filter(
            id_klasifikasi_jenis_data__nama_klasifikasi='LANGKA'
        ).count()

        row = {
            'kategori_ilap': ilap.id_kategori_ilap.nama_kategori_ilap if ilap.id_kategori_ilap else '',
            'nama_ilap': ilap.nama_ilap,
            'jenis_data_wajib': wajib_count,
            'jenis_data_penting': penting_count,
            'jenis_data_lengkap': lengkap_count,
            'jenis_data_langka': langka_count,
        }
        data.append(row)

    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
@user_passes_test(is_pmde_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def laporan_rekap_himpun_olah_data_export(request):
    """Export Laporan Rekap Penghimpunan dan Pengolahan Data to XLSX or PDF."""
    params = request.GET if request.method == 'GET' else request.POST
    export_format = params.get('format', 'excel')

    # Get filtered ILAPs
    ilaps = _get_filtered_data(params)

    if export_format == 'excel':
        return _export_to_excel(ilaps)
    elif export_format == 'pdf':
        return _export_to_pdf(ilaps)
    else:
        return HttpResponse('Format not supported', status=400)


def _export_to_excel(ilaps):
    """Export data to Excel format."""
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Himpun & Olah Data"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Write title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'REKAP PENGHIMPUNAN DAN PENGOLAHAN DATA'
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = center_alignment

    # Write subtitle
    ws.merge_cells('A2:G2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = 'Informasi IPC Penghimpunan Data yang telah disampaikan ke AP'
    subtitle_cell.font = Font(italic=True, size=10)
    subtitle_cell.alignment = center_alignment

    # Write headers
    headers = ['NO', 'KATEGORI ILAP', 'NAMA ILAP', 'JENIS DATA WAJIB', 'JENIS DATA PENTING', 'JENIS DATA LENGKAP', 'JENIS DATA LANGKA']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Write data
    for row_idx, ilap in enumerate(ilaps, start=5):
        # Get jenis data for this ILAP
        jenis_data_list = ilap.jenis_data_ilap.all()

        # Count by klasifikasi
        wajib_count = jenis_data_list.filter(
            id_klasifikasi_jenis_data__nama_klasifikasi='WAJIB'
        ).count()
        penting_count = jenis_data_list.filter(
            id_klasifikasi_jenis_data__nama_klasifikasi='PENTING'
        ).count()
        lengkap_count = jenis_data_list.filter(
            id_klasifikasi_jenis_data__nama_klasifikasi='LENGKAP'
        ).count()
        langka_count = jenis_data_list.filter(
            id_klasifikasi_jenis_data__nama_klasifikasi='LANGKA'
        ).count()

        row_data = [
            row_idx - 4,
            ilap.id_kategori_ilap.nama_kategori_ilap if ilap.id_kategori_ilap else '',
            ilap.nama_ilap,
            wajib_count,
            penting_count,
            lengkap_count,
            langka_count,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx > 3:
                cell.alignment = center_alignment

    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_data = excel_file.getvalue()

    # Create response
    filename = f"Rekap_Penghimpunan_Pengolahan_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_to_pdf(ilaps):
    """Export data to PDF format."""
    # For now, we'll return a placeholder
    # In production, use reportlab or similar
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        # Create PDF
        buffer = BytesIO()
        pdf = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)

        # Prepare data
        data = [['NO', 'KATEGORI ILAP', 'NAMA ILAP', 'JENIS DATA WAJIB', 'JENIS DATA PENTING', 'JENIS DATA LENGKAP', 'JENIS DATA LANGKA']]

        for row_idx, ilap in enumerate(ilaps, start=1):
            jenis_data_list = ilap.jenis_data_ilap.all()
            wajib_count = jenis_data_list.filter(id_klasifikasi_jenis_data__nama_klasifikasi='WAJIB').count()
            penting_count = jenis_data_list.filter(id_klasifikasi_jenis_data__nama_klasifikasi='PENTING').count()
            lengkap_count = jenis_data_list.filter(id_klasifikasi_jenis_data__nama_klasifikasi='LENGKAP').count()
            langka_count = jenis_data_list.filter(id_klasifikasi_jenis_data__nama_klasifikasi='LANGKA').count()

            data.append([
                str(row_idx),
                ilap.id_kategori_ilap.nama_kategori_ilap if ilap.id_kategori_ilap else '',
                ilap.nama_ilap,
                str(wajib_count),
                str(penting_count),
                str(lengkap_count),
                str(langka_count),
            ])

        # Create table
        table = Table(data, colWidths=[0.6*inch, 1.5*inch, 1.8*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])

        # Add table style
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        # Build PDF
        story = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=6,
            alignment=TA_CENTER
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=12,
            alignment=TA_CENTER
        )

        story.append(Paragraph('REKAP PENGHIMPUNAN DAN PENGOLAHAN DATA', title_style))
        story.append(Paragraph('Informasi IPC Penghimpunan Data yang telah disampaikan ke AP', subtitle_style))
        story.append(table)

        pdf.build(story)

        # Return response
        filename = f"Rekap_Penghimpunan_Pengolahan_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except ImportError:
        # If reportlab not installed, return Excel instead
        return _export_to_excel(ilaps)
