"""Laporan Register Penerimaan Data view - Data Receipt Register Report."""

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.http import require_GET, require_http_methods
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_protect
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..models.tiket import Tiket
from ..models.detil_tanda_terima import DetilTandaTerima
from ..utils import format_periode


def _is_p3de_user(user):
    """Check if user is P3DE user or admin."""
    return user.is_superuser or user.is_staff or user.groups.filter(
        name__in=['user_p3de', 'admin', 'admin_p3de']
    ).exists()


class LaporanRegisterPenerimaanView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Display Data Receipt Register Report filtered by month and year.

    Template: laporan_register_penerimaan/list.html
    """
    template_name = 'laporan_register_penerimaan/list.html'

    def test_func(self):
        return _is_p3de_user(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_year = datetime.now().year
        years = list(range(current_year, current_year - 10, -1))
        context['years'] = years
        context['current_year'] = current_year
        context['current_month'] = datetime.now().month
        return context


@login_required
@user_passes_test(_is_p3de_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def register_penerimaan_data(request):
    """DataTables server-side endpoint for Register Penerimaan Data.

    Filters tikets by tgl_terima_dip within the specified month and year.

    Parameters:
    - bulan: Month (1-12)
    - tahun: Year
    - draw, start, length: DataTables pagination

    Returns JSON with tiket data including tanda terima info.
    """
    params = request.POST if request.method == 'POST' else request.GET
    bulan = params.get('bulan')
    tahun = params.get('tahun')

    try:
        draw = int(params.get('draw', 1))
    except (ValueError, TypeError):
        draw = 1

    try:
        start = int(params.get('start', 0))
    except (ValueError, TypeError):
        start = 0

    try:
        length = int(params.get('length', 25))
    except (ValueError, TypeError):
        length = 25

    if not bulan or not tahun:
        return JsonResponse({'draw': draw, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': []})

    try:
        bulan = int(bulan)
        tahun = int(tahun)
        if bulan < 1 or bulan > 12:
            return JsonResponse({'draw': draw, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': []})
    except (ValueError, TypeError):
        return JsonResponse({'draw': draw, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': []})

    start_date = datetime(tahun, bulan, 1).date()
    if bulan == 12:
        end_date = datetime(tahun + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(tahun, bulan + 1, 1).date() - timedelta(days=1)

    tikets = Tiket.objects.filter(
        tgl_terima_dip__date__gte=start_date,
        tgl_terima_dip__date__lte=end_date,
    ).select_related(
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap',
        'id_periode_data__id_periode_pengiriman',
        'id_bentuk_data',
    ).prefetch_related(
        'detiltandaterima_set__id_tanda_terima',
        'id_periode_data__id_sub_jenis_data_ilap__klasifikasijenisdata_set__id_klasifikasi_tabel',
    ).order_by('tgl_terima_dip', 'id')

    records_total = tikets.count()
    records_filtered = records_total

    tikets_page = tikets[start:start + length]

    # Pre-fetch tanda terima for these tikets
    tiket_ids = [t.id for t in tikets_page]
    detil_qs = DetilTandaTerima.objects.filter(
        id_tiket__in=tiket_ids
    ).select_related('id_tanda_terima').order_by('id_tanda_terima__tanggal_tanda_terima')

    tanda_terima_map = {}
    for detil in detil_qs:
        tanda_terima_map.setdefault(detil.id_tiket_id, []).append(detil.id_tanda_terima)

    data = []
    for idx, tiket in enumerate(tikets_page, start=start + 1):
        sub_jenis = tiket.id_periode_data.id_sub_jenis_data_ilap
        ilap = sub_jenis.id_ilap

        # Dasar hukum via KlasifikasiJenisData
        dasar_hukum_list = [
            k.id_klasifikasi_tabel.deskripsi
            for k in sub_jenis.klasifikasijenisdata_set.all()
        ]
        dasar_hukum = ', '.join(dasar_hukum_list) if dasar_hukum_list else '-'

        # Tanda terima
        tanda_terima_entries = tanda_terima_map.get(tiket.id, [])
        if tanda_terima_entries:
            tt = tanda_terima_entries[0]
            nomor_tt = tt.nomor_tanda_terima_format
            tanggal_tt = tt.tanggal_tanda_terima.strftime('%d/%m/%Y') if tt.tanggal_tanda_terima else '-'
        else:
            nomor_tt = '-'
            tanggal_tt = '-'

        periode_pengiriman = tiket.id_periode_data.id_periode_pengiriman
        deskripsi_periode = periode_pengiriman.periode_penerimaan if periode_pengiriman else ''

        data.append({
            'no': idx,
            'nama_ilap': ilap.nama_ilap,
            'dasar_hukum': dasar_hukum,
            'jenis_data_ilap': sub_jenis.nama_jenis_data,
            'bentuk_data': tiket.id_bentuk_data.deskripsi if tiket.id_bentuk_data else '-',
            'periode_tahun_data': format_periode(deskripsi_periode, tiket.periode, tiket.tahun),
            'jumlah_data_diterima': tiket.baris_diterima or 0,
            'nomor_surat_pengantar': tiket.nomor_surat_pengantar or '-',
            'tanggal_surat_pengantar': tiket.tanggal_surat_pengantar.strftime('%d/%m/%Y') if tiket.tanggal_surat_pengantar else '-',
            'nomor_tanda_terima': nomor_tt,
            'tanggal_tanda_terima': tanggal_tt,
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data,
    })


@login_required
@user_passes_test(_is_p3de_user)
@require_GET
def register_penerimaan_export(request):
    """Export Register Penerimaan Data to XLSX.

    GET Parameters:
    - bulan: Month (1-12)
    - tahun: Year
    """
    bulan = request.GET.get('bulan')
    tahun = request.GET.get('tahun')

    if not bulan or not tahun:
        return HttpResponse('Invalid parameters', status=400)

    try:
        bulan = int(bulan)
        tahun = int(tahun)
        if bulan < 1 or bulan > 12:
            return HttpResponse('Invalid month', status=400)
    except (ValueError, TypeError):
        return HttpResponse('Invalid parameters', status=400)

    bulan_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                   'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    periode_label = f'{bulan_names[bulan]} {tahun}'

    start_date = datetime(tahun, bulan, 1).date()
    if bulan == 12:
        end_date = datetime(tahun + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(tahun, bulan + 1, 1).date() - timedelta(days=1)

    tikets = Tiket.objects.filter(
        tgl_terima_dip__date__gte=start_date,
        tgl_terima_dip__date__lte=end_date,
    ).select_related(
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap',
        'id_periode_data__id_periode_pengiriman',
        'id_bentuk_data',
    ).prefetch_related(
        'id_periode_data__id_sub_jenis_data_ilap__klasifikasijenisdata_set__id_klasifikasi_tabel',
    ).order_by('tgl_terima_dip', 'id')

    tiket_ids = [t.id for t in tikets]
    detil_qs = DetilTandaTerima.objects.filter(
        id_tiket__in=tiket_ids
    ).select_related('id_tanda_terima').order_by('id_tanda_terima__tanggal_tanda_terima')

    tanda_terima_map = {}
    for detil in detil_qs:
        tanda_terima_map.setdefault(detil.id_tiket_id, []).append(detil.id_tanda_terima)

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Register Penerimaan Data'

    # Title row
    ws.merge_cells('A1:K1')
    ws['A1'] = f'Register Penerimaan Data - {periode_label}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header row
    headers = [
        'No', 'Nama ILAP', 'Dasar Hukum', 'Jenis Data ILAP', 'Bentuk Data',
        'Periode / Tahun Data', 'Jumlah Data Diterima',
        'Nomor Surat Pengantar', 'Tanggal Surat Pengantar',
        'Nomor Tanda Terima', 'Tanggal Tanda Terima',
    ]
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row_idx, tiket in enumerate(tikets, 3):
        idx = row_idx - 2
        sub_jenis = tiket.id_periode_data.id_sub_jenis_data_ilap
        ilap = sub_jenis.id_ilap
        dasar_hukum_list = [k.id_klasifikasi_tabel.deskripsi for k in sub_jenis.klasifikasijenisdata_set.all()]
        dasar_hukum = ', '.join(dasar_hukum_list) if dasar_hukum_list else '-'

        tanda_terima_entries = tanda_terima_map.get(tiket.id, [])
        if tanda_terima_entries:
            tt = tanda_terima_entries[0]
            nomor_tt = tt.nomor_tanda_terima_format
            tanggal_tt = tt.tanggal_tanda_terima.strftime('%d/%m/%Y') if tt.tanggal_tanda_terima else '-'
        else:
            nomor_tt = '-'
            tanggal_tt = '-'

        periode_pengiriman = tiket.id_periode_data.id_periode_pengiriman
        deskripsi_periode = periode_pengiriman.periode_penerimaan if periode_pengiriman else ''

        row_data = [
            idx,
            ilap.nama_ilap,
            dasar_hukum,
            sub_jenis.nama_jenis_data,
            tiket.id_bentuk_data.deskripsi if tiket.id_bentuk_data else '-',
            format_periode(deskripsi_periode, tiket.periode, tiket.tahun),
            tiket.baris_diterima or 0,
            tiket.nomor_surat_pengantar or '-',
            tiket.tanggal_surat_pengantar.strftime('%d/%m/%Y') if tiket.tanggal_surat_pengantar else '-',
            nomor_tt,
            tanggal_tt,
        ]

        alt_fill = PatternFill(start_color='DEEAF1', end_color='DEEAF1', fill_type='solid') if idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if alt_fill:
                cell.fill = alt_fill

    # Auto-fit column widths
    col_widths = [5, 30, 25, 30, 15, 18, 20, 25, 22, 30, 22]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[2].height = 30

    excel_file = BytesIO()
    wb.save(excel_file)

    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Register_Penerimaan_Data_{periode_label.replace(" ", "_")}.xlsx"'
    )
    return response
