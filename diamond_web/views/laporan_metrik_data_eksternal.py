"""Laporan Metrik Data Eksternal view - Third Party Data (External Data) Metric Report."""

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.http import require_GET, require_http_methods
from django.views.generic import TemplateView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_protect
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook

from ..models.tiket import Tiket
from ..forms.laporan_metrik_data_eksternal import LaporanMetrikDataEksternalFilterForm, LaporanMetrikDataEksternalExportResource
from ..constants.jenis_tabel import JENIS_TABEL_DIIDENTIFIKASI, JENIS_TABEL_TIDAK_DIIDENTIFIKASI


def _is_pide_user(user):
    """Check if user is PIDE user or admin."""
    return user.is_superuser or user.is_staff or user.groups.filter(name__in=['user_pide', 'admin', 'admin_pide']).exists()


def _get_filtered_tikets(params):
    """Utility function to filter tikets based on request parameters."""
    tgl_mulai_str = params.get('tgl_mulai')
    tgl_akhir_str = params.get('tgl_akhir')
    id_ilap = params.get('id_ilap')
    id_jenis_data = params.get('id_jenis_data')
    nama_sub_jenis_data = params.get('nama_sub_jenis_data')
    nama_tabel_I = params.get('nama_tabel_I')

    # Query base
    tikets = Tiket.objects.all().select_related(
        'id_periode_data__id_sub_jenis_data_ilap__id_ilap',
        'id_periode_data__id_sub_jenis_data_ilap__id_jenis_tabel'
    )
    
    # Filter by date range
    if tgl_mulai_str:
        try:
            # datetime-local format: YYYY-MM-DDTHH:MM
            tgl_mulai = datetime.strptime(tgl_mulai_str, '%Y-%m-%dT%H:%M')
            tikets = tikets.filter(tgl_transfer__gte=tgl_mulai)
        except ValueError:
            pass
            
    if tgl_akhir_str:
        try:
            tgl_akhir = datetime.strptime(tgl_akhir_str, '%Y-%m-%dT%H:%M')
            tikets = tikets.filter(tgl_transfer__lte=tgl_akhir)
        except ValueError:
            pass
            
    # Filter by ILAP
    if id_ilap and id_ilap != 'all' and id_ilap != '':
        tikets = tikets.filter(id_periode_data__id_sub_jenis_data_ilap__id_ilap_id=id_ilap)
        
    # Filter by Jenis Data
    if id_jenis_data and id_jenis_data != 'all' and id_jenis_data != '':
        tikets = tikets.filter(id_periode_data__id_sub_jenis_data_ilap_id=id_jenis_data)
        
    # Filter by Subjenis Data
    if nama_sub_jenis_data and nama_sub_jenis_data != 'all' and nama_sub_jenis_data != '':
        tikets = tikets.filter(id_periode_data__id_sub_jenis_data_ilap__nama_sub_jenis_data=nama_sub_jenis_data)
        
    # Filter by Tabel I
    if nama_tabel_I and nama_tabel_I != 'all' and nama_tabel_I != '':
        tikets = tikets.filter(id_periode_data__id_sub_jenis_data_ilap__nama_tabel_I=nama_tabel_I)

    # Order by transfer date (chronological)
    return tikets.order_by('tgl_transfer')


class LaporanMetrikDataEksternalView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Display Laporan Metrik Data Eksternal by filtering tikets
    based on a date range and other parameters.
    """
    template_name = 'laporan_metrik_data_eksternal/list.html'
    
    def test_func(self):
        """Verify user is PIDE user or admin."""
        return _is_pide_user(self.request.user)
    
    def get_context_data(self, **kwargs):
        """Add form to context."""
        context = super().get_context_data(**kwargs)
        context['form'] = LaporanMetrikDataEksternalFilterForm()
        return context


@login_required
@user_passes_test(_is_pide_user)
@require_http_methods(["GET", "POST"])
@csrf_protect
def laporan_metrik_data_eksternal_data(request):
    """DataTables server-side endpoint for Third Party Data (External Data) Metric Report.
    
    Filters tikets by:
    - tgl_transfer within specified tgl_mulai and tgl_akhir
    - id_ilap, id_jenis_data, nama_sub_jenis_data, nama_tabel_I
    
    Returns JSON with tiket data.
    """
    params = request.POST if request.method == 'POST' else request.GET
    
    try:
        draw = int(params.get('draw', 1))
        start = int(params.get('start', 0))
        length = int(params.get('length', 10))
    except (ValueError, TypeError):
        draw, start, length = 1, 0, 10
    
    # Get filtered tikets using helper
    tikets = _get_filtered_tikets(params)
    
    records_total = Tiket.objects.count()
    records_filtered = tikets.count()
    
    # Pagination
    tikets_paginated = tikets[start:start + length]
    
    # Build response data
    data = []
    for tiket in tikets_paginated:
        sub_jenis_data = tiket.id_periode_data.id_sub_jenis_data_ilap if tiket.id_periode_data else None
        if not sub_jenis_data:
            continue
            
        ilap = sub_jenis_data.id_ilap
        
        # Logic for calculated fields (matching LaporanTransferExportResource)
        baris_diterima = tiket.baris_diterima or 0
        baris_u = tiket.baris_u or 0
        baris_i = tiket.baris_i or 0
        
        # Diidentifikasi (id=1), Tidak Diidentifikasi (id=2)
        id_jenis_tabel = sub_jenis_data.id_jenis_tabel_id
        
        data_teridentifikasi_i = baris_i if id_jenis_tabel == JENIS_TABEL_DIIDENTIFIKASI else 0
        data_tidak_diidentifikasi_i = baris_i if id_jenis_tabel == JENIS_TABEL_TIDAK_DIIDENTIFIKASI else 0
        
        persentase_identifikasi = ''
        if baris_diterima > 0:
            persentase_identifikasi = f"{(baris_i / baris_diterima * 100):.2f}%"
        
        row = {
            'nama_ilap': ilap.nama_ilap if ilap else '',
            'nama_jenis_data': sub_jenis_data.nama_jenis_data,
            'nama_sub_jenis_data': sub_jenis_data.nama_sub_jenis_data,
            'nama_tabel_I': sub_jenis_data.nama_tabel_I,
            'nomor_tiket': tiket.nomor_tiket,
            'data_diterima': baris_diterima,
            'data_teridentifikasi_i': data_teridentifikasi_i,
            'data_tidak_teridentifikasi_u': baris_u,
            'data_tidak_diidentifikasi_i': data_tidak_diidentifikasi_i,
            'data_res': tiket.baris_res,
            'persentase_identifikasi': persentase_identifikasi
        }
        data.append(row)
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data
    })


@login_required
@user_passes_test(_is_pide_user)
@require_GET
@csrf_protect
def laporan_metrik_data_eksternal_export(request):
    """Export Laporan Metrik Data Eksternal to XLSX file."""
    # Get filtered tikets using helper
    tikets = _get_filtered_tikets(request.GET)
    
    # Use LaporanMetrikDataEksternalExportResource
    resource = LaporanMetrikDataEksternalExportResource()
    dataset = resource.export(tikets)
    
    # Create Excel workbook using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Transfer"
    
    # Write headers
    headers = dataset.headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data rows
    for row_idx, row in enumerate(dataset, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_data = excel_file.getvalue()
    
    # Create filename
    filename = "Laporan_Transfer"
    tgl_mulai_str = request.GET.get('tgl_mulai')
    tgl_akhir_str = request.GET.get('tgl_akhir')
    if tgl_mulai_str and tgl_akhir_str:
        filename += f"_{tgl_mulai_str[:10]}_ke_{tgl_akhir_str[:10]}"
    
    # Create HTTP response
    response = HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response

